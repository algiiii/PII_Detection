"""Generate a CNIL-format ROPA sample workbook (ODS + XLSX) with Italian content,
then run the B1 ingestion + dictionary mapping on it and print the result.

The sample mirrors the **full layout** of the real CNIL template's activity sheet
(tab ``4_-_Example_``): every section is present — stakeholders, purposes, data
subjects, recipients, security measures, transfers — but only the "Categories of
personal data" block is filled with data that matters to us, exactly as a real
register looks. The generator proves the normalizer picks out *only* that block
and skips the rest, across several activities and both file formats.

The structural labels stay in English (the normalizer keys off them, as in the
CNIL English template); the categories are filled in Italian, as an Italian
company using the template would. One category carries a **cell comment** (the
CNIL "red triangle" note) to exercise the annotation-stripping in the ODS reader.

Output goes to ``corpus/ropa/`` (gitignored); import ``cnil_sample.ods`` from the
web review UI (``/ropa`` → import) to watch the register populate.

Run from the project root:
    ./.venv/bin/python examples/ropa/demo.py
"""

from __future__ import annotations

import argparse
import tempfile
from pathlib import Path

from odf.office import Annotation
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import Workbook
from openpyxl.comments import Comment

from pii_detection.ropa.ingestion.category_mapper import (
    CategoryMapper,
    build_dictionary_mapper,
    build_llm_category_mapper,
)
from pii_detection.ropa.ingestion.pipeline import ingest_file, map_categories
from pii_detection.ropa.repository import ROPARepository


def _mapper(kind: str) -> CategoryMapper:
    """Build the mapper selected on the command line (dictionary/llm/hybrid)."""
    if kind == "dictionary":
        return build_dictionary_mapper()
    if kind == "llm":
        return build_llm_category_mapper(use_fallback=False)
    return build_llm_category_mapper(use_fallback=True)


HERE = Path(__file__).resolve().parent

# A cell is either a plain string or a (value, comment) pair — the comment becomes
# a cell annotation (ODS) / comment (XLSX), i.e. a CNIL "red triangle" note.
Cell = str | tuple[str, str]

# A category is (macro label, description, retention[, comment on the macro cell]).
Category = tuple[str, str, str] | tuple[str, str, str, str]

# (sheet, activity name, purpose, sub-purpose, data subjects, [categories]).
ACTIVITIES: list[tuple[str, str, str, str, str, list[Category]]] = [
    (
        "Gestione_HR",
        "Gestione del personale",
        "Amministrazione del rapporto di lavoro",
        "Calcolo delle retribuzioni",
        "Dipendenti",
        [
            ("Stato civile, identità, dati identificativi", "nome, cognome e indirizzo di residenza", "5 anni"),
            ("Informazioni economiche e finanziarie", "coordinate bancarie", "10 anni"),
            ("Contatti", "indirizzo email e numero di telefono", "a criterio"),
            (
                "Numero di previdenza sociale (o codice fiscale)",
                "codice fiscale",
                "5 anni",
                "Cfr. art. 87 del Regolamento: regole nazionali specifiche per questo dato.",
            ),
        ],
    ),
    (
        "Marketing",
        "Newsletter e marketing",
        "Invio di comunicazioni commerciali",
        "Gestione dei consensi",
        "Clienti e potenziali clienti",
        [
            ("Contatti", "email", "24 mesi"),
            ("Stato civile, identità, dati identificativi", "dati anagrafici", "24 mesi"),
            ("Profilazione", "preferenze di acquisto profilate", "a criterio"),
        ],
    ),
    (
        "Videosorveglianza",
        "Videosorveglianza sede",
        "Sicurezza dei locali",
        "Protezione dei beni aziendali",
        "Dipendenti e visitatori",
        [
            ("Immagini", "immagini delle telecamere", "7 giorni"),
            ("Dati di connessione", "indirizzo IP del dispositivo", "6 mesi"),
        ],
    ),
]


def _sheet_rows(
    name: str, purpose: str, sub_purpose: str, subjects: str, categories: list[Category]
) -> list[list[Cell]]:
    """Build one activity sheet mirroring the full CNIL template layout.

    Only the "Categories of personal data" block carries the data we ingest; every
    other section is present with placeholder content so the sheet looks like a
    real CNIL fiche and the normalizer has to skip it.

    :param name: activity name (the ``Name of the processing operation`` value).
    :param purpose: the ``Main purpose`` value.
    :param sub_purpose: a secondary purpose (ignored on ingestion).
    :param subjects: the categories of data subjects (ignored on ingestion).
    :param categories: the declared data categories, as
        ``(macro, description, retention[, comment])`` tuples.
    :returns: the sheet grid, cells being plain strings or ``(value, comment)`` pairs.
    """
    rows: list[list[Cell]] = [
        ["Description of the processing operation"],
        ["Name of the processing operation", name],
        ["N° / REF", name],
        ["Data of creation of the processing", "May 26, 2018"],
        ["Update of the processing", "May 13, 2019"],
        [],
        ["Stakeholders", "Name", "Address", "ZIP Code", "Town", "Country", "Phone", "Email"],
        ["Controller", "Louise DUPONT", "1 rue Rivoli", "75001", "Paris", "France", "01 xx", "c@ex.com"],
        ["Data protection officer", "Martin HENRI", "1 rue Rivoli", "75001", "Paris", "France", "01 xx", "dpo@ex.com"],
        ["DPO's Organisation (if external DPO)", "N/A"],
        [],
        ["Purpose(s) of the data processing"],
        ["Main purpose", purpose],
        ["Sub-purpose 1", sub_purpose],
        [],
        ["Categories of personal data", "Description", "Data retention period"],
    ]
    for category in categories:
        macro, description, retention = category[0], category[1], category[2]
        macro_cell: Cell = (macro, category[3]) if len(category) == 4 else macro
        rows.append([macro_cell, description, retention])
    rows.extend(
        [
            [],
            ["Categories of data subjects", "Description", "Details"],
            ["Category 1", subjects],
            [],
            ["Recipients", "Type of recipient", "Details"],
            ["Recipient 1", "Internal department that processes the data", "Administrative Department"],
            [],
            ["Security measures", "Type of security measure", "Details"],
            ["Security measure 1", "User access control"],
            [],
            ["Transfers to third countries or international organisations", "Recipient", "Country"],
            ["Recipient organisation 1", "N/A", "N/A"],
        ]
    )
    return rows


def _xlsx_cell(value: Cell) -> str:
    """Return the plain text of a cell (the comment, if any, is dropped for XLSX)."""
    return value[0] if isinstance(value, tuple) else value


def write_xlsx(path: Path) -> None:
    """Write the sample workbook as ``.xlsx`` (openpyxl).

    :param path: destination ``.xlsx`` path (overwritten if present).
    """
    wb = Workbook()
    default = wb.active
    for sheet, name, purpose, sub_purpose, subjects, categories in ACTIVITIES:
        ws = wb.create_sheet(sheet)
        for row in _sheet_rows(name, purpose, sub_purpose, subjects, categories):
            ws.append([_xlsx_cell(cell) for cell in row])
            for column, cell in enumerate(row, start=1):
                if isinstance(cell, tuple):
                    ws.cell(row=ws.max_row, column=column).comment = Comment(cell[1], "CNIL")
    tutorial = wb.create_sheet("Tutorial", 0)
    tutorial.append(["Record of processing activities — one sheet per processing operation."])
    if default is not None:
        wb.remove(default)
    wb.save(path)


def _ods_row(cells: list[Cell]) -> TableRow:
    """Build one ODS table row; a ``(value, comment)`` cell gets an annotation.

    :param cells: the row's cells, plain strings or ``(value, comment)`` pairs.
    :returns: the populated :class:`odf.table.TableRow`.
    """
    row = TableRow()
    for value in cells:
        # office:value-type="string" so every viewer renders the text as the cell
        # value (without it some tools show the cell empty), like the real CNIL file.
        cell = TableCell(valuetype="string")
        if isinstance(value, tuple):
            text, comment = value
            annotation = Annotation()
            annotation.addElement(P(text=comment))
            cell.addElement(annotation)  # the CNIL "red triangle" note
        else:
            text = value
        cell.addElement(P(text=text))
        row.addElement(cell)
    return row


def write_ods(path: Path) -> None:
    """Write the sample workbook as ``.ods`` (odfpy).

    :param path: destination ``.ods`` path (overwritten if present).
    """
    doc = OpenDocumentSpreadsheet()
    tutorial = Table(name="Tutorial")
    tutorial.addElement(
        _ods_row(["Record of processing activities — one sheet per processing operation."])
    )
    doc.spreadsheet.addElement(tutorial)
    for sheet, name, purpose, sub_purpose, subjects, categories in ACTIVITIES:
        table = Table(name=sheet)
        for row in _sheet_rows(name, purpose, sub_purpose, subjects, categories):
            table.addElement(_ods_row(row))
        doc.spreadsheet.addElement(table)
    doc.save(str(path))


def show(path: Path, kind: str) -> None:
    """Ingest a workbook, map its categories with the chosen mapper, print the tree.

    :param path: the workbook to ingest.
    :param kind: the mapper to run (``dictionary``/``llm``/``hybrid``).
    """
    db_url = f"sqlite:///{tempfile.mkdtemp()}/ropa.db"
    activities = ingest_file(path, db_url)
    map_categories(ROPARepository(db_url), _mapper(kind))
    print(f"\n=== {path.name} [{kind}]: {len(activities)} trattamenti (Tutorial saltato) ===")
    for activity in ROPARepository(db_url).load():
        print(f"[{activity.id}] {activity.name}")
        for macro in activity.macro_categories:
            print(f"   macro {macro.raw_text!r}  retention_months={macro.retention_months}")
            for category in macro.categories:
                print(f"      - {category.raw_text!r:44} -> {category.pii_types}")


def main() -> None:
    """Generate both workbooks into ``corpus/ropa/`` and run the demo on each."""
    parser = argparse.ArgumentParser(description="Generate a CNIL-format ROPA sample and map it.")
    parser.add_argument(
        "--mapper",
        choices=["dictionary", "llm", "hybrid"],
        default="dictionary",
        help="which category mapper to run (default: dictionary)",
    )
    args = parser.parse_args()

    target = HERE.parents[1] / "corpus" / "ropa"  # repo-root/corpus/ropa (gitignored)
    target.mkdir(parents=True, exist_ok=True)
    ods, xlsx = target / "cnil_sample.ods", target / "cnil_sample.xlsx"
    write_ods(ods)
    write_xlsx(xlsx)
    print(f"generati:\n  {ods}\n  {xlsx}")
    show(ods, args.mapper)
    show(xlsx, args.mapper)


if __name__ == "__main__":
    main()
