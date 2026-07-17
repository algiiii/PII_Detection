"""Read a tabular ROPA spreadsheet into raw, still-uninterpreted records.

Block B1, ingestion step. This module only *extracts* the sheet content; it does
no interpretation (splitting, type resolution, ``pii_type`` mapping): that is the
job of :mod:`pii_detection.ropa.ingestion.normalizer`.

Current layout: ``per_row`` — the first row holds the column names and every
following row is one processing activity. The ``activity_id`` is intentionally
absent from the sheet: it is generated downstream by the normalizer.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class RawTable:
    """The spreadsheet content, read but not yet interpreted.

    :ivar columns: the column names, i.e. the header row, in sheet order.
    :ivar records: one dict per activity row, keyed by column name; values are
        raw cell values (``str``, ``int`` or ``None`` for empty cells).
    """

    columns: tuple[str, ...]
    records: list[dict[str, Any]]


def read_records(path: str | Path) -> RawTable:
    """Read the active worksheet of an ``.xlsx`` ROPA file.

    :param path: path to the spreadsheet.
    :returns: the header and the row records as a :class:`RawTable`.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = next(rows)
        records = [dict(zip(header, row, strict=False)) for row in rows]
    finally:
        wb.close()
    return RawTable(columns=tuple(header), records=records)
