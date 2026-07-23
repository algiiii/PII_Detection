"""Read spreadsheet sheets (ODS or Excel) into plain string grids — block B1.

The reader is format-agnostic on the outside: :func:`sheet_names` lists the tabs
of a workbook and :func:`read_sheet` returns one tab as a ``list[list[str]]``
grid, dispatching on the file extension (``.ods`` via odfpy, ``.xlsx``/``.xlsm``
via openpyxl). Keeping both back-ends behind the same shape lets the normalizer
stay unaware of the source format (Adapter).
"""

from pathlib import Path
from openpyxl import load_workbook
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableCell
from odf.text import P


def sheet_names(path: str | Path) -> list[str]:
    """List the sheet (tab) names of a workbook, in document order.

    Lets the ingestion iterate every tab and decide per-sheet whether it is a
    processing-activity record, without hard-coding tab names.

    :param path: path to the ``.ods`` or ``.xlsx``/``.xlsm`` workbook.
    :returns: the sheet names, in the order they appear in the workbook.
    :raises ValueError: if the file extension is not a supported spreadsheet.
    """
    suffix = Path(path).suffix.lower()
    if suffix == ".ods":
        doc = load_ods(path)
        return [str(t.getAttribute("name")) for t in doc.spreadsheet.getElementsByType(Table)]
    if suffix in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    raise ValueError(f"unsupported spreadsheet format: {suffix!r}")


def read_sheet(path: str | Path, sheet_name: str) -> list[list[str]]:
    """Read one sheet of a spreadsheet into a grid of cell strings.

    Dispatches on the file extension: ``.ods`` via odfpy, ``.xlsx``/``.xlsm``
    via openpyxl. Both back-ends yield the same shape, so the normalizer is
    format-agnostic.

    :param path: path to the ``.ods`` or ``.xlsx``/``.xlsm`` workbook.
    :param sheet_name: name of the sheet to read.
    :returns: the sheet as a grid of stripped cell strings (trailing empty cells
        of each row are dropped).
    :raises ValueError: if the file extension is not a supported spreadsheet.
    :raises KeyError: if no sheet with ``sheet_name`` exists.
    """
    suffix = Path(path).suffix.lower()
    if suffix == ".ods":
        return _read_ods(path, sheet_name)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx(path, sheet_name)
    raise ValueError(f"unsupported spreadsheet format: {suffix!r}")


def _read_xlsx(path: str | Path, sheet_name: str) -> list[list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [("" if v is None else str(v)).strip() for v in row]
            while cells and cells[-1] == "":
                cells.pop()
            rows.append(cells)
        return rows
    finally:
        wb.close()


def _read_ods(path: str | Path, sheet_name: str) -> list[list[str]]:
    doc = load_ods(path)
    for table in doc.spreadsheet.getElementsByType(Table):
        if table.getAttribute("name") != sheet_name:
            continue
        rows = []
        for tr in table.getElementsByType(TableRow):
            cells = []
            for tc in tr.getElementsByType(TableCell):
                rep = int(tc.getAttribute("numbercolumnsrepeated") or 1)
                text = "".join(str(p) for p in tc.getElementsByType(P)).strip()
                cells.extend([text] * min(rep, 40))
            while cells and cells[-1] == "":
                cells.pop()
            rows.append(cells)
        return rows
    raise KeyError(f"sheet not found: {sheet_name!r}")